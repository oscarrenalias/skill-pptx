"""Write operations: cell mutations with atomic save."""

from __future__ import annotations

import os
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, Optional, Union


def _infer_type(value: str) -> Any:
    """Infer Python type from a string value.

    Rules (in order):
        value.startswith("=")  →  str (formula string, kept as-is)
        int(value) succeeds    →  int
        float(value) succeeds  →  float
        else                   →  str
    """
    if value.startswith("="):
        return value
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        pass
    return value


def set_cell(
    file: Union[str, Path],
    sheet: str,
    cell: str,
    value: str,
) -> Dict[str, Any]:
    """Set a single cell to an inferred-type value and save atomically.

    Type inference rules (applied to the raw string *value*):
        - Starts with ``=``       → stored as a formula string
        - Parseable as ``int``    → stored as int
        - Parseable as ``float``  → stored as float
        - Otherwise               → stored as str

    The file is written to a temp file in the same directory as *file*, then
    renamed over the original with ``os.replace()`` so the write is atomic.

    Args:
        file:  Path to the .xlsx file.
        sheet: Name of the worksheet.
        cell:  Cell address in A1 notation (e.g. ``"B3"``).
        value: The string value to write (type-inferred before storing).

    Returns:
        dict with keys:
            sheet (str): The sheet name as given.
            cell (str):  The cell address, normalised to uppercase.
            value:       The stored value after type inference.
    """
    import openpyxl  # local import to keep the module importable without openpyxl installed

    path = Path(file)
    try:
        wb = openpyxl.load_workbook(str(path), data_only=False)
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)

    if sheet not in wb.sheetnames:
        print(f"error: sheet '{sheet}' not found", file=sys.stderr)
        wb.close()
        sys.exit(1)

    ws = wb[sheet]
    cell_upper = cell.upper()
    typed_value = _infer_type(value)
    ws[cell_upper] = typed_value

    # Atomic save: write to a temp file in the same directory, then replace.
    dir_path = path.parent
    try:
        fd, tmp_path = tempfile.mkstemp(dir=str(dir_path), suffix=".tmp")
        os.close(fd)
        wb.save(tmp_path)
        os.replace(tmp_path, str(path))
    except Exception as exc:
        # Clean up temp file on failure.
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)

    return {
        "sheet": sheet,
        "cell": cell_upper,
        "value": typed_value,
    }


def add_sheet(
    file: Union[str, Path],
    name: str,
    position: Optional[int] = None,
) -> Dict[str, Any]:
    """Insert a new blank sheet into the workbook and save atomically.

    Args:
        file:     Path to the .xlsx file.
        name:     Name for the new sheet.
        position: 1-based insertion position.  Default (``None``) appends at end.

    Returns:
        dict with keys:
            name (str):     The new sheet name.
            position (int): The 1-based position of the new sheet after insertion.
    """
    import openpyxl  # local import to keep the module importable without openpyxl installed

    path = Path(file)
    try:
        wb = openpyxl.load_workbook(str(path), data_only=False)
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)

    if name in wb.sheetnames:
        print(f"error: sheet '{name}' already exists", file=sys.stderr)
        wb.close()
        sys.exit(1)

    # openpyxl.create_sheet uses a 0-based index; None appends at end.
    insert_index: Optional[int] = None
    if position is not None:
        insert_index = position - 1  # convert 1-based to 0-based

    wb.create_sheet(title=name, index=insert_index)
    actual_position = wb.sheetnames.index(name) + 1  # 1-based

    dir_path = path.parent
    try:
        fd, tmp_path = tempfile.mkstemp(dir=str(dir_path), suffix=".tmp")
        os.close(fd)
        wb.save(tmp_path)
        os.replace(tmp_path, str(path))
    except Exception as exc:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)

    return {
        "name": name,
        "position": actual_position,
    }


def delete_sheet(
    file: Union[str, Path],
    name: str,
) -> Dict[str, Any]:
    """Delete a sheet by name and save atomically.

    Args:
        file: Path to the .xlsx file.
        name: Name of the sheet to delete.

    Returns:
        dict with keys:
            deleted (str): The name of the deleted sheet.
    """
    import openpyxl  # local import to keep the module importable without openpyxl installed

    path = Path(file)
    try:
        wb = openpyxl.load_workbook(str(path), data_only=False)
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)

    if name not in wb.sheetnames:
        print(f"error: sheet '{name}' not found", file=sys.stderr)
        wb.close()
        sys.exit(1)

    if len(wb.sheetnames) == 1:
        print("error: cannot delete the last sheet", file=sys.stderr)
        wb.close()
        sys.exit(1)

    del wb[name]

    dir_path = path.parent
    try:
        fd, tmp_path = tempfile.mkstemp(dir=str(dir_path), suffix=".tmp")
        os.close(fd)
        wb.save(tmp_path)
        os.replace(tmp_path, str(path))
    except Exception as exc:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)

    return {"deleted": name}


def rename_sheet(
    file: Union[str, Path],
    old_name: str,
    new_name: str,
) -> Dict[str, Any]:
    """Rename a sheet and save atomically.

    Args:
        file:     Path to the .xlsx file.
        old_name: Current name of the sheet.
        new_name: New name for the sheet.

    Returns:
        dict with keys:
            old_name (str): The original sheet name.
            new_name (str): The new sheet name.
    """
    import openpyxl  # local import to keep the module importable without openpyxl installed

    path = Path(file)
    try:
        wb = openpyxl.load_workbook(str(path), data_only=False)
    except Exception as exc:
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)

    if old_name not in wb.sheetnames:
        print(f"error: sheet '{old_name}' not found", file=sys.stderr)
        wb.close()
        sys.exit(1)

    if new_name in wb.sheetnames:
        print(f"error: sheet '{new_name}' already exists", file=sys.stderr)
        wb.close()
        sys.exit(1)

    wb[old_name].title = new_name

    dir_path = path.parent
    try:
        fd, tmp_path = tempfile.mkstemp(dir=str(dir_path), suffix=".tmp")
        os.close(fd)
        wb.save(tmp_path)
        os.replace(tmp_path, str(path))
    except Exception as exc:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        print(f"error: {exc}", file=sys.stderr)
        sys.exit(1)

    return {"old_name": old_name, "new_name": new_name}
